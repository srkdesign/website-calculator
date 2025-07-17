import pandas as pd
from openpyxl.styles import Font
from apply_formatting import format_to_currency, adjust_columns

currency_formats = {
    "USD": '"$"#,##0.00',
    "RUB": "#,##0.00 ₽",
}


# Exporting data to an excel sheet
def export_to_excel(
    all_pages_data,
    section_bp,
    page_bp,
    currency_code,
    currency_formats=currency_formats,
):
    rows = []
    outline_levels = []

    grand_total = 0.0
    total_page_cost = 0.0

    for page_data in all_pages_data:
        page_name = page_data["name_field"].value
        total_page_cost = 0.0

        page_rows = []
        page_outline_levels = []

        # Subrows = Sections on the Page
        for row in page_data["subrows"]:
            title = row["title"].value
            desc = row["description"].value
            try:
                difficulty = float(row["difficulty"].value)
            except:
                difficulty = 0.0
            price = float(section_bp.value) * difficulty
            total_page_cost += price

            page_rows.append(
                {
                    "Страница": "",
                    "Блок": title,
                    "Описание": desc,
                    "Сложность": difficulty,
                    "Стоимость": price,
                }
            )
            page_outline_levels.append(1)

        total_page_cost = max(total_page_cost, float(page_bp.value))
        grand_total += total_page_cost

        # Page-level row
        page_rows.insert(
            0,
            {
                "Страница": page_name,
                "Блок": "",
                "Описание": "",
                "Сложность": "",
                "Стоимость": total_page_cost,
            },
        )
        page_outline_levels.insert(0, 0)

        rows.extend(page_rows)
        outline_levels.extend(page_outline_levels)

    df = pd.DataFrame(rows)

    df.loc[len(df.index)] = {
        "Страница": "",
        "Блок": "",
        "Описание": "",
        "Сложность": "Итого",
        "Стоимость": grand_total,
    }

    filename = "project_estimate.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Расчет стоимости сайта")
        ws = writer.sheets["Расчет стоимости сайта"]

        for i, level in enumerate(outline_levels, start=2):
            row_dim = ws.row_dimensions[i]
            row_dim.outlineLevel = level
            row_dim.collapsed = level == 1

        ws.sheet_properties.outlinePr.summaryBelow = True

        for cell in ws[1]:
            cell.font = Font(bold=True)

        adjust_columns(ws)
        format_to_currency(ws, currency_code, currency_formats)

        last_row = ws.max_row
        ws[f"D{last_row}"].font = Font(bold=True)
        ws[f"E{last_row}"].font = Font(bold=True)

    print(f"✅ Excel file saved: {filename}")
