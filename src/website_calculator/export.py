import pandas as pd
from openpyxl.styles import Font, PatternFill
from website_calculator.apply_formatting import format_to_currency, adjust_columns

currency_formats = {
  "USD": '"$" #,##0.00',       # Excel-compliant format
  "RUB": '#,##0.00" ₽"',      # Ruble with space and quotes
}

# Exporting data to an excel sheet
def export_to_excel(filename, pages, section_bp_field, page_bp_field, currency_code, currency_formats=currency_formats):
  print(f"📦 [DEBUG] currency_code = {currency_code}")
  try:
    headers_ru = ["Страница", "Блок", "Описание", "Сложность", "Стоимость"]
    headers_en = ["Page", "Section", "Description", "Difficulty", "Cost"]
    headers = headers_en if currency_code == "USD" else headers_ru

    total_label = "Total" if currency_code == "USD" else "Итого"
    cost_column_name = headers[-1]

    rows = []
    outline_levels = []

    grand_total = 0.0

    section_bp = section_bp_field()
    page_bp = page_bp_field()

    for page in pages:
      page_name = page.title.text()
      total_page_cost = 0.0

      page_rows = []
      page_outline_levels = []

      for section in page.sections:
        title = section.title.text()
        desc = section.desc.text()
        difficulty = section.get_difficulty()
        price = section_bp * difficulty
        total_page_cost += price

        page_rows.append({
          headers[0]: "",
          headers[1]: title,
          headers[2]: desc,
          headers[3]: difficulty,
          headers[4]: price,
        })
        page_outline_levels.append(1)

      total_page_cost = max(total_page_cost, page_bp)
      grand_total += total_page_cost

      page_rows.insert(0, {
        headers[0]: page_name,
        headers[1]: "",
        headers[2]: "",
        headers[3]: "",
        headers[4]: total_page_cost,
      })
      page_outline_levels.insert(0, 0)

      rows.extend(page_rows)
      outline_levels.extend(page_outline_levels)

    print(f"🧪 [DEBUG] Using headers: {headers}")
    df = pd.DataFrame(rows)
    df.loc[len(df.index)] = {
      headers[0]: "",
      headers[1]: "",
      headers[2]: "",
      headers[3]: total_label,
      headers[4]: grand_total,
    }

    print(f"[DEBUG] Saving Excel to: {filename}")

    sheet_name = "Website Quote" if currency_code == "USD" else "Расчет стоимости сайта"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
      df.to_excel(writer, index=False, sheet_name=sheet_name)
      ws = writer.sheets["Website Quote"] if currency_code == "USD" else writer.sheets["Расчет стоимости сайта"]

      for i, level in enumerate(outline_levels, start=2):
        row_dim = ws.row_dimensions[i]
        row_dim.outlineLevel = level
        row_dim.collapsed = level == 1

      ws.sheet_properties.outlinePr.summaryBelow = True

      for cell in ws[1]:
        cell.font = Font(bold=True)

      adjust_columns(ws)
      format_to_currency(ws, currency_code, currency_formats, target_column_name=cost_column_name)

      header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
      for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

      last_row = ws.max_row
      ws[f"D{last_row}"].font = Font(bold=True)
      ws[f"E{last_row}"].font = Font(bold=True)

    print(f"✅ [DEBUG] Excel file saved: {filename}")

  except Exception as e:
      print("[DEBUG] Failed to export Excel file:", e)
