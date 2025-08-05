from openpyxl.utils import get_column_letter

def format_to_currency(
    ws, currency_code, currency_formats, target_column_name="Стоимость"
):
  target_col = None
  for idx, cell in enumerate(ws[1], start=1):
    if cell.value == target_column_name:
      target_col = get_column_letter(idx)
      break

  if not target_col:
    print(f"⚠️ Column '{target_column_name}' not found.")
    return

  fmt = currency_formats.get(currency_code, "#,##0.00")

  for row in range(2, ws.max_row + 1):
    cell = ws[f"{target_col}{row}"]

    # Ensure cell value is a number
    try:
      cell.value = float(cell.value)
      cell.number_format = fmt
    except (TypeError, ValueError):
      pass  # Skip non-numeric cells


def adjust_columns(ws):
  for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)

    for cell in col:
      if cell.value:
        cell_length = len(str(cell.value))
        if cell_length > max_length:
          max_length = cell_length
    
    adjusted_width = max_length + 5
    ws.column_dimensions[col_letter].width = adjusted_width